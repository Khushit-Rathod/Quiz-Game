from tkinter import *
def w0():
    global rt,lb
    rt=Tk()    #creates a window
    rt.title("QUIZ GAME")
    rt.geometry("1500x900")
    from PIL import Image,ImageTk
    image=Image.open(r'C:\\Users\\Prafull\\Downloads\\quiz.jpeg')
    resized_image=image.resize((1000,750),Image.ANTIALIAS)
    photo=ImageTk.PhotoImage(resized_image)
    lb=Label(rt,image=photo,borderwidth=5,highlightbackground="black",highlightthickness=12)
    lb.pack()
    lb.after(4000,w1)
    rt.state('zoomed')
    rt.mainloop()
 
def w1():
    rt.destroy()
    global root
    root=Tk()
    root.title("QUIZ GAME")
    root.geometry("1500x900")
    from PIL import Image,ImageTk
    image=Image.open(r"C:\\Users\\Prafull\\Downloads\\dddd.jpeg")
    resized_image = image.resize((1500,900),Image.ANTIALIAS)
    photo=ImageTk.PhotoImage(resized_image)
    lbl=Label(root,image=photo)
    lbl.pack()
    def tab1():
        from time import strftime
        def myt():
            time_string=strftime("%H:%M:%S%p \n %A \n %x")
            l1.config(text=time_string,borderwidth=4,highlightbackground="black",highlightthickness=3)
            l1.after(1000,myt)
        l1=Label(root,font=("garamond",20,"bold"),bg='white')
        l1.place(x=1325,y=10)
        myt()
        global label_t,label1,label2,entry1,entry2,button1,label_q,label_o,button_o,button_p,button_l,button_k
        label_q=Label(root,text="  WELCOME TO THE QUIZ GAME  ",font=('garamond',50,'bold','italic'),fg='white',bg='black',borderwidth=4,highlightbackground="white",highlightthickness=5)
        label_q.place(x=185,y=50)
        label_o=Label(root,text="  Please click any one option  ",font=('garamond',40,'bold'),borderwidth=5,highlightbackground="black",highlightthickness=5)
        label_o.place(x=450,y=175)
        button_o=Button(root,text="  START A NEW QUIZ ",font=('arial',20,'bold'),command=tab2,borderwidth=9,highlightbackground="black",highlightthickness=7,bg="white")
        button_o.place(x=600,y=300)
        button_p=Button(root,text="      INSTRUCTIONS    ",font=('arial',20,'bold'),command=instructions,borderwidth=9,highlightbackground="black",highlightthickness=7,bg="white")
        button_p.place(x=600,y=400)
        button_l=Button(root,text="     LEADERBOARD     ",font=('arial',20,'bold'),command=leaderboard,borderwidth=9,highlightbackground="black",highlightthickness=7,bg="white")
        button_l.place(x=600,y=500)
        button_k=Button(root,text="              EXIT              ",font=('arial',20,'bold'),command=lambda :root.destroy(),borderwidth=9,highlightbackground="black",highlightthickness=7,bg="white")
        button_k.place(x=600,y=600)
 
    def instructions():
        global label_t,label1,label2,entry1,entry2,button1,label_q,label_o,button_o,button_p,button_l,button_k
        label_q.destroy()
        label_o.destroy()
        button_o.destroy()
        button_p.destroy()
        button_l.destroy()
        button_k.destroy()
        label_1=Label(root,text="INSTRUCTIONS",font=('garamond',50),fg='black',borderwidth=5,highlightbackground="black",highlightthickness=5)
        label_1.place(x=500,y=35)
        label_1=Label(root,text="1. This test consist of 2 rounds.",font=('garamond',30),fg='black')
        label_1.place(x=175,y=150)
        label_1=Label(root,text="2. Each round has 10 multiple choice questions.",font=('garamond',30),fg='black')
        label_1.place(x=175,y=225)
        label_1=Label(root,text="3. There will be no choice to go to the previous question,\nhence once answered that will be considered Final.",font=('garamond',30),fg='black')
        label_1.place(x=175,y=300)
        label_1=Label(root,text="4. There will be no negative marking.",font=('garamond',30),fg='black')
        label_1.place(x=175,y=425)
        label_1=Label(root,text="5. You cannot skip any question.",font=('garamond',30),fg='black')
        label_1.place(x=175,y=500)
        label_1=Label(root,text="6. If you answer more than 6  Questions in Round 1,\nthen only you can go to next Round.",font=('garamond',30),fg='black')
        label_1.place(x=175,y=575)
        b1=Button(root,text="START QUIZ",font=('arial',20),command=w2)
        b1.place(x=625,y=725)
 
    def tab2():
        global user_name,em_ail,label_t,label1,label2,entry1,entry2,button1,label_q,label_o,button_o,button_p,button_l,button_k,user_name,em_ail
        user_name = StringVar() 
        em_ail = StringVar() 
        label_q.destroy()
        label_o.destroy()
        button_o.destroy()
        button_p.destroy()
        button_l.destroy()
        button_k.destroy()
        label_t=Label(root,text=" NEW QUIZ ",font=('garamond',50),fg='black',borderwidth=2,highlightbackground="black",highlightthickness=3)
        label_t.place(x=575,y=50)
        label1=Label(root,text=" Username :",font=('arial',30),fg='black',borderwidth=2,highlightbackground="black",highlightthickness=3)
        label1.place(x=300,y=250)
        label2=Label(root,text=" Email ID :   ",font=('arial',30),fg='black',borderwidth=2,highlightbackground="black",highlightthickness=3)
        label2.place(x=300,y=350)
        entry1=Entry(root,font=('arial',30),textvariable=user_name)
        entry1.place(x=550,y=250)
        entry2=Entry(root,font=('arial',30),textvariable=em_ail)
        entry2.place(x=550,y=350)
        button1=Button(root,text='SUBMIT',font=('arial',30),command=check,borderwidth=9,highlightbackground="black",highlightthickness=7)
        button1.place(x=675,y=500)
        root.bind('<Return>',check1)
   
    def check1(e):
        global ename
        ename=em_ail.get()
        if "@" in ename:
            op=Toplevel(root)
            op.title("Quiz Game")
            op.geometry("500x350")
            op_label=Label(op,text="Proceed to QUIZ ?",font=('garamond',30,'bold'))
            op_label.pack(pady=20)
            mb=Button(op,text="Confirm",command=w2,font=('arial',18,'bold'))
            mb.place(x=215,y=100) 
        else:
            pop=Toplevel(root)
            pop.title("ERROR")
            pop.geometry("500x350")
            pop_label=Label(pop,text="Invalid Email-id , Please re-enter",font=('garamond',18,'bold'))
            pop_label.pack(pady=20)
            mb=Button(pop,text="OK",font=('arial',12,'bold'),command=lambda :[tab2(),pop.destroy()])
            mb.place(x=225,y=100)
    
    def check():
        global ename
        ename=em_ail.get()
        if "@" in ename:
            op=Toplevel(root)
            op.title("Quiz Game")
            op.geometry("500x350")
            op_label=Label(op,text="Proceed to QUIZ ?",font=('garamond',30,'bold'))
            op_label.pack(pady=20)
            mb=Button(op,text="Confirm",command=w2,font=('arial',18,'bold'))
            mb.place(x=215,y=100)      
        else:
            pop=Toplevel(root)
            pop.title("ERROR")
            pop.geometry("500x350")
            pop_label=Label(pop,text="Invalid Email-id , Please re-enter",font=('garamond',18,'bold'))
            pop_label.pack(pady=20)
            mb=Button(pop,text="OK",font=('arial',12,'bold'),command=lambda :[tab2(),pop.destroy()])
            mb.place(x=225,y=100)
    tab1()
    root.state('zoomed')
    root.mainloop()
 
def w2():
    root.destroy()
    global question_no,score,total_no_questions,option1,option2,option3,option4,question,entry1,entry2,uname,ename
    global win,user_name,em_ail
    uname = user_name.get()
    ename = em_ail.get()
    win=Tk()
    win.title("QUIZ GAME")
    win.geometry("1500x900")
    from PIL import Image,ImageTk
    image=Image.open(r"C:\\Users\\Prafull\\Downloads\\slide1.jpeg")
    resized_image=image.resize((1360,760),Image.ANTIALIAS)
    photo=ImageTk.PhotoImage(resized_image)
    lb2=Label(win,image=photo,borderwidth=5,highlightbackground="black",highlightthickness=7)
    lb2.pack()
    questions=[ "1. If (16(2/3)%) of a number is added to itself, \n \t the number becomes 1400. Find the original number.",
        "2. If the radius of a circle is increased by 50%,\t \n  \t  what will be the percentage increase in the area?",
        "3. A mother is twice as old as her son. If 20 years ago,\n the age of the mother was 10 times the age of the son, what is \n the present age of the mother?",
        "4. A shopkeeper sold an article for Rs. 2500.\n  If the cost price of the article is 2000, find the profit percent.",
        "\t 5. What is the area of a triangle with base \t \t \n5 meters and height 10 meters?",
        "          6. Predict the element in the sequence -  1,8,27,_,125.     \n ",
        " \t 7. What is the probability of getting a prime\t \t \n  number when a dice is rolled?",
        " \t 8. Look at this series: 12, 11, 13, 12, 14, 13, …\t\t \n What number should come next?",
        "\t \t  9. Find the odd man out.\t \t \t ",
        "             10. Which word does NOT belong with the other?     \t"]
 
    options=[["1100","1356","1000","1200"],
        ["150%","50%","100%","125%"],
        ["38 years","40 years","43 years","45 years"],
        ["23%","25%","27%","30%"],
        ["25 square meters","35 square meters","20 square meters","40 square meters"],
        ["50","64","56","36"],
        ["1/2","1/3","1/6","2/3"],
        ["13","14","16","15"],
        ["Index","Acknowledgment","Chapter","Book"],
        ["Inch","Metre","Ounce","Yard"]]
 
    ans=[4,4,3,2,1,2,1,4,4,3]
    def check(option):
        if(option==1):
            val2.set(0)
            val3.set(0)
            val4.set(0)
        if(option==2):
            val1.set(0)
            val3.set(0)
            val4.set(0)
        if(option==3):
            val2.set(0)
            val1.set(0)
            val4.set(0)
        if(option==4):
            val2.set(0)
            val3.set(0)
            val1.set(0)
    score=0
    total_no_questions=10
    question_no=1
 
    a1=Label(win,text="Quiz Round 1 Begins!!",font=('garamond',40,'bold'),fg='black',bg="#d9c181",borderwidth=5,highlightbackground="black",highlightthickness=3)
    a1.place(x=475,y=70)
    question=Label(win,text=questions[0],font=('garamond',30,"bold"),bg="black",fg="white",justify="center")
    question.place(x=225,y=175)
   
 
    val1=IntVar()
    val2=IntVar()
    val3=IntVar()
    val4=IntVar()
 
    option1=Radiobutton(win,variable=val1,value=1,text=options[0][0],command=lambda :check(1),font=('garamond',20,"bold"),bg="#d9c181")
    option1.place(x=675,y=325)
    option2=Radiobutton(win,variable=val2,value=1,text=options[0][1],command=lambda :check(2),font=('garamond',20,"bold"),bg="#d9c181")
    option2.place(x=675,y=400)
    option3=Radiobutton(win,variable=val3,value=1,text=options[0][2],command=lambda :check(3),font=('garamond',20,"bold"),bg="#d9c181")
    option3.place(x=675,y=475)
    option4=Radiobutton(win,variable=val4,value=1,text=options[0][3],command=lambda :check(4),font=('garamond',20,"bold"),bg="#d9c181")
    option4.place(x=675,y=550)
 
   
 
    def next():
        global question_no,score,total_no_questions,option1,option2,option3,option4,question,bt1,lb1,bt2,pu
        a1.destroy()
        if(val1.get()==1):
            selected_option=1
        elif(val2.get()==1):
            selected_option=2
        elif(val3.get()==1):
            selected_option=3
        elif(val4.get()==1):
            selected_option=4
        else:
            selected_option=-1
        if(ans[question_no-1]==selected_option):
            score+=1
        question_no+=1
        if(question_no>total_no_questions):
            n1.destroy()
            option1.destroy()
            option2.destroy()
            option3.destroy()
            option4.destroy()
            question.destroy()
            win.destroy()
            pu=Tk()
            pu.title("End of round 1")
            pu.geometry("1500x900")
            pu.configure(bg="#F2503B")
            s1=Label(pu,text="\t SCORE = "+str(score)+"                 ",font=('garamond',30,'bold'),bg="#FEBE28",borderwidth=2,highlightbackground="black",highlightthickness=3)
            s1.place(relx=.33,rely=.36)
            s2=Label(pu,text="            END OF ROUND 1    \t",font=('garamond',50,'bold'),fg='black',borderwidth=2,highlightbackground="black",highlightthickness=3)
            s2.place(x=300,y=135)
            if(score>=4):
                lb2=Label(pu,text="\t       YOU ARE QUALIFIED FOR THE NEXT ROUND\t\t     ",font=('garamond',20,'bold'),borderwidth=2,highlightbackground="black",highlightthickness=3)
                lb2.place(x=300,y=480)
                bt1=Button(pu,text="NEXT ROUND",font=('garamond',30,'bold'),command=w3,bg="#FEBE28",borderwidth=9,highlightbackground="black",highlightthickness=7)
                bt1.place(x=615,y=600)
            else:
               
                lb1=Label(pu,text="\tYOU ARE NOT QUALIFIED FOR THE NEXT ROUND\t",font=('garamond',25,'bold'),borderwidth=2,highlightbackground="black",highlightthickness=3)
                lb1.place(x=200,y=450)
                bt2=Button(pu,text=" QUIT ",font=('garamond',28,'bold'),bg="#FEBE28",command=lambda :pu.destroy(),borderwidth=9,highlightbackground="black",highlightthickness=7)
                bt2.place(x=705,y=600)
            pu.mainloop()
        else:
            val1.set(0)
            val2.set(0)
            val3.set(0)
            val4.set(0)
            question.config(text=questions[question_no-1])
            option1.config(text=options[question_no-1][0])
            option2.config(text=options[question_no-1][1])
            option3.config(text=options[question_no-1][2])
            option4.config(text=options[question_no-1][3])
 
    n1=Button(win,text="NEXT",font=('arial',30,'bold'),bg="#b48e3d",command=next,borderwidth=9,highlightbackground="black",highlightthickness=7)
    n1.place(x=650,y=625)
    va1=Label(win,text="Username: "+uname,font=('arial',20,'bold'),fg='black',bg="#d9c181")
    va1.place(x=1100,y=70)
    win.state('zoomed')
    win.mainloop()
 
def w3():
    pu.destroy()
    global ques,sc,total_questions,opt1,opt2,opt3,opt4,question,bt1,uname,ename
    global win1,user_name,em_ail
    uname = user_name.get()
    ename = em_ail.get()
    win1=Tk()
    win1.title("QUIZ GAME")
    win1.geometry("1500x900")
    from PIL import Image,ImageTk
    image=Image.open(r"C:\\Users\\Prafull\\Downloads\\slide1.jpeg")
    resized_image=image.resize((1360,760),Image.ANTIALIAS)
    photo=ImageTk.PhotoImage(resized_image)
    lb2=Label(win1,image=photo)
    lb2.pack()
 
    questions=[ "\t 1. Python files are saved with the extension as.....\t",
        " \t \t      2. Userinput is read as .....\t \t ",
        "\t 3. Single Line Comments in Python begin with .....\t\t",
        "\t\t4. If I  ______ a Bird I would Fly \t\t",
        "\t\t5. Chirag hardly ever cooks , ___ ?\t\t",
        "\t6. We have been in Sector 12 , Navi Mumbai ____ 1999.",
        "7. Which of the following operators is the correct option for power(ab)?",
        "\t8. What will be the output of the following statements?\n X=’23’ + ‘16’ \nprint(X)",
        "\t9. What will be the output of the following statements?\nX=[‘today’ , ‘tomorrow’ , ‘hi’ ]\nprint(X[1])",
        "\t10. What will be the output of the following statements?\nA = “world”\nprint(a[::-1])"]
 
    options=[[".python",".pe       ",".py       ",".pi        "],
        ["Floating Decimal","Text String","Boolean Value","Integer"],
        ["{  ","%  ","*  ","#  "],
        ["is  ","am ","was  ","were"],
        ["isn't he","does he","he doesn't","doesn't he"],
        ["for","from","about","since"],
        ["a^b","a*b","a**b","a^*b"],
        ["38","39","2316","error"],
        ["hi","tomorrow","1","today"],
        ["world","-1","dlrow","error"]]
 
    ans=[3,2,4,4,2,4,3,3,4,3]
 
    def check(option):
        if(option==1):
            val2.set(0)
            val3.set(0)
            val4.set(0)
        if(option==2):
            val1.set(0)
            val3.set(0)
            val4.set(0)
        if(option==3):
            val2.set(0)
            val1.set(0)
            val4.set(0)
        if(option==4):
            val2.set(0)
            val3.set(0)
            val1.set(0)
    sc=0
    total_questions=10
    ques=1
 
    aa=Label(win1,text="Quiz Round 2 Begins!!",font=('garamond',50,'bold'),fg='black',bg="#d9c181",borderwidth=5,highlightbackground="black",highlightthickness=3)
    aa.place(x=425,y=60)
    question=Label(win1,text=questions[0],font=('garamond',28,"bold"),bg="black",fg="white",justify="center")
    question.place(x=200,y=180)
 
    val1=IntVar()
    val2=IntVar()
    val3=IntVar()
    val4=IntVar()
 
    opt1=Radiobutton(win1,variable=val1,value=1,text=options[0][0],command=lambda :check(1),font=('garamond',20,"bold"),bg="#d9c181")
    opt1.place(x=675,y=350)
    opt2=Radiobutton(win1,variable=val2,value=1,text=options[0][1],command=lambda :check(2),font=('garamond',20,"bold"),bg="#d9c181")
    opt2.place(x=675,y=425)
    opt3=Radiobutton(win1,variable=val3,value=1,text=options[0][2],command=lambda :check(3),font=('garamond',20,"bold"),bg="#d9c181")
    opt3.place(x=675,y=500)
    opt4=Radiobutton(win1,variable=val4,value=1,text=options[0][3],command=lambda :check(4),font=('garamond',20,"bold"),bg="#d9c181")
    opt4.place(x=675,y=575)
 
    def next():
        global ques,sc,total_questions,opt1,opt2,opt3,opt4,question,o
        aa.destroy()
        if(val1.get()==1):
            selected_option=1
        elif(val2.get()==1):
            selected_option=2
        elif(val3.get()==1):
            selected_option=3
        elif(val4.get()==1):
            selected_option=4
        else:
            selected_option=-1
        if(ans[ques-1]==selected_option):
            sc+=1
        ques+=1
        if(ques>total_questions):
            n1.destroy()
            opt1.destroy()
            opt2.destroy()
            opt3.destroy()
            opt4.destroy()
            question.destroy()
            o=int(score+sc)
            s2=Label(win1,text="   END OF ROUND 2 \n SCORE = "+str(sc)+" ",font=('garamond',40,'bold'),bg="#d9c181",borderwidth=5,highlightbackground="black",highlightthickness=5)
            s2.place(x=510,y=235)
            s2.after(5000,rrr)
            win1.mainloop()
 
        else:
            val1.set(0)
            val2.set(0)
            val3.set(0)
            val4.set(0)
            question.config(text=questions[ques-1])
            opt1.config(text=options[ques-1][0])
            opt2.config(text=options[ques-1][1])
            opt3.config(text=options[ques-1][2])
            opt4.config(text=options[ques-1][3])
 
    n1=Button(win1,text="   NEXT   ",font=('arial',25,'bold'),bg="#b48e3d",command=next,borderwidth=9,highlightbackground="black",highlightthickness=7)
    n1.place(x=640,y=650)
    va2=Label(win1,text="Username: "+uname,font=('arial',20,'bold'),fg='black',bg="#d9c181")
    va2.place(x=1100,y=60)
    win1.state('zoomed')
    win1.mainloop()
def email():
    import smtplib  
    global o,ename,uname
    s = smtplib.SMTP('smtp.gmail.com', 587)
    s.starttls()
    s.login("manali.22110360@viit.ac.in","wecandoit")
    message = "Hello "+uname +", Below is the score of the game that you played \n Score="+str(o)
    recipients = ['manalij873@gmail.com','patilpushkar189@gmail.com','suyogmanke4@gmail.com','chaitanya.22110114@viit.ac.in','khushit.22110024@viit.ac.in']
    recipients.append(ename)
    print(recipients)
    s.sendmail("manali.22110360@viit.ac.in", recipients, message)
    s.quit()
 
def rrr():
    global o,wi
    win1.destroy()
    wi=Tk()
    wi.title("QUIZ GAME")
    wi.geometry("1500x900")
    from PIL import Image,ImageTk
    image=Image.open(r"C:\\Users\\Prafull\\Downloads\\ffff.jpeg")
    resized_image=image.resize((1450,750),Image.ANTIALIAS)
    photo=ImageTk.PhotoImage(resized_image)
    dd=Label(wi,image=photo,borderwidth=5,highlightbackground="black",highlightthickness=12)
    dd.pack()
    pe=Label(wi,text=" THANK YOU FOR PLAYING!! ",font=('garamond',40,'bold'),borderwidth=3,highlightbackground="black",highlightthickness=3)
    pe.place(x=400,y=100)
    pb=Button(wi,text=" Quit ",font=('garamond',20,'bold'),command=lambda :wi.destroy(),borderwidth=9,highlightbackground="black",highlightthickness=7)
    pb.place(x=600,y=600)
    pc=Button(wi,text=" Send my score ",font=('garamond',20,'bold'),command=lambda :[save_data(),email(),access()],borderwidth=9,highlightbackground="black",highlightthickness=7)
    pc.place(x=800,y=600)
    pd=Label(wi,text=" TOTAL SCORE = "+str(o)+" ",font=('garamond',45,'bold'),borderwidth=3,highlightbackground="black",highlightthickness=3)
    pd.place(x=485,y=375)
    def access():
        eee=Toplevel(wi)
        eee.title("Leaderboard")
        eee.geometry("500x350")
        op_label=Label(eee,text="Your score has been sent...\n Do you want to see the Leaderboard",font=('garamond',18,'bold'))
        op_label.pack(pady=20)
        mb=Button(eee,text="OK",command=lambda :[leaderboard(),eee.destroy()],font=('arial',12,'bold'))
        mb.place(x=225,y=100)
    wi.state('zoomed')
    wi.mainloop()
 
def leaderboard():
    from tkinter import ttk
    import pandas as pd
    global a,df
    df=pd.read_excel(r'C:\\Users\\Prafull\\Desktop\\Quiz_Data.xlsx')
    a=df.sort_values(by=['Score','Username'],ascending=[False,True])
    print(a)
    ram = Tk() #creates a window
    ram.geometry("1500x900")
    style = ttk.Style()
    style.theme_use('clam')
    style.configure("Treeview")
    frame = Frame(ram)
    frame.pack(padx=200,pady=100)
    tree = ttk.Treeview(frame)
    tree["column"] = list(a.columns)
    tree["show"] = "headings"
    for col in tree["column"]:
        tree.heading(col, text=col)
    df_rows = a.to_numpy().tolist()
    for row in df_rows:
        tree.insert("", "end", values=row)
        tree.pack()
    ram.state('zoomed')
    ram.mainloop()
 
def save_data():
    global uname,ename,user_name,em_ail,o,score,sc,date
    from openpyxl import load_workbook
    from time import strftime
    date=strftime("%x")
    uname=user_name.get()
    ename=em_ail.get()
    wb= load_workbook(r'C:\\Users\\Prafull\\Desktop\\Quiz_Data.xlsx')
    ws=wb.active   #can make changes(add or delete)
    ws.title="DATA"
    ws['A1']="Played on"
    ws['B1']="Username"
    ws['C1']="Email-id"
    ws['D1']="Score"
    ws.append([date,uname,ename,o])
    wb.save(r'C:\\Users\\Prafull\\Desktop\\Quiz_Data.xlsx')
w0()
